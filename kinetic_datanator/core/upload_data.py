from kinetic_datanator.data_source import refseq
import os
from os import path
from Bio import SeqIO
import openpyxl
from kinetic_datanator.core import common_schema, models
import json
import sqlalchemy_utils
import sqlalchemy


CACHE_DIRNAME = os.path.join(os.path.dirname(__file__), '..', 'data_source', 'cache')

class Uploader():
    def __init__(self, parent_node, dict_of_new_objects):
        self.parent_node = parent_node
        self.dict_of_new_objects = dict_of_new_objects

class UploadData():
    def __init__(self, cache_dirname=CACHE_DIRNAME):
        self.cache_dirname = cache_dirname
        self.flask = common_schema.CommonSchema(cache_dirname=cache_dirname)


    def upload_reference_genome(self, path_to_annotation_file):
        #bio_seqio_object = SeqIO.parse(path_to_annotation_file, "genbank")
        refseq.Refseq(cache_dirname=self.cache_dirname).load_content([SeqIO.parse(path_to_annotation_file, "genbank")])


    def upload_rna_rseq_experiment(self, path_to_template_directory):
        print(path_to_template_directory)
        filename = "{}/RNA-SeqMetadataTemplate.xlsx".format(path_to_template_directory)
        wb = openpyxl.load_workbook(filename=filename)
        ws = wb.get_sheet_by_name('Experiments')
        for i in range(2, ws.max_row+1):
            experiment_name = wb.get_sheet_by_name('Experiments').cell(row=i,column=1).value
            experiment_description = wb.get_sheet_by_name('Experiments').cell(row=i,column=2).value
            new_experiment = self.flask.get_or_create_object(models.RNASeqExperiment,
                exp_name=experiment_name,
                accession_number=experiment_name)
            print(experiment_name)
            print(experiment_description)
            self.flask.session.add(new_experiment)
        self.flask.session.commit()

    def get_or_create_object_upload(self, session, cls, kwargs):
        """ Get the first instance of :obj:`cls` that has the property-values pairs described by kwargs, or create an instance of :obj:`cls`
        if there is no instance with the property-values pairs described by kwargs
        Args:
            cls (:obj:`class`): type of object to find or create
            **kwargs: values of the properties of the object
        Returns:
            :obj:`Base`: instance of :obj:`cls` hat has the property-values pairs described by kwargs
        """
        q = session.query(cls).filter_by(**kwargs)
        if session.query(q.exists()).scalar():
            return q.first()

        obj = cls(**kwargs)
        session.add(obj)
        return obj

    def get_primitive_fields(self, object, object_type):
        print(object_type)
        list_of_primitives = {}
        for field in object:
            if (type(object[field]) is not list) and (type(object[field]) is not dict):
                list_of_primitives[field] = object[field]
            if type(object[field]) is dict:
                new_type = sqlalchemy_utils.get_type(object_type.__dict__[field])
                new_object = self.upload_from_json(object[field], new_type)
                list_of_primitives[field] = new_object

        return list_of_primitives

    def get_dict_of_new_objects(self, object):
        dict_of_new_objects = {}
        for field in object:
            if type(object[field]) is list:
                dict_of_new_objects[field] = object[field]
        return dict_of_new_objects




    def upload_from_json(self, objects, first_data_type):
        session = self.flask.session

        if type(first_data_type) is str:
            models_object = models.__dict__[first_data_type]
        else:
            models_object = first_data_type

        data_type = models_object

        to_upload = []
        dict_of_new_objects = {}
        list_of_fields = self.get_primitive_fields(objects, data_type)
        dict_of_new_objects = self.get_dict_of_new_objects(objects)
        
        first_parent = self.get_or_create_object_upload(session, models_object, list_of_fields)
        first_upload = Uploader(parent_node=first_parent, dict_of_new_objects=dict_of_new_objects)
        to_upload.append(first_upload)

        while to_upload:
            new_to_upload = []
            for uploader in to_upload:
                for key in uploader.dict_of_new_objects:
                    new_obj_type = sqlalchemy_utils.get_type(type(uploader.parent_node).__dict__[key])
                    list_of_new_objects = uploader.dict_of_new_objects[key]
                    for new_object in list_of_new_objects:
                        the_new_object =self.get_or_create_object_upload(session, new_obj_type, self.get_primitive_fields(new_object, new_obj_type))
                        getattr(uploader.parent_node, key).append(the_new_object)
                        dict_of_new_objects = self.get_dict_of_new_objects(new_object)
                        if dict_of_new_objects:
                            new_uploader = Uploader(parent_node = the_new_object, dict_of_new_objects = dict_of_new_objects)
                            new_to_upload.append(new_uploader)
            to_upload = new_to_upload
        session.commit()
        return first_parent

        

        """

    def upload_processed_data(sample_name, csv_file, top_directory=path.dirname(__file__)):
        #this takes in a csv file, and puts the file in the proper place so that it can be queried. 
        new_pandas = pd.read_csv(csv_file, sep=',').set_index("gene_locus")
        new_pandas.to_pickle("{}/{}".format(top_directory, sample_name))

        """


