""" 
:Author: Yosef Roth <yosefdroth@gmail.com>
:Date: 2017-04-06
:Copyright: 2017, Karr Lab
:License: MIT
"""

from . import taxon_finder
from openpyxl import Workbook
import os


def create_excel_sheet(species, reactions, output_filename):
    """ Save a list of reaction kinetic data to an Excel workbook

    Args:
        species (:obj:`str`): species
        reactions (:obj:`list` of :obj:`kinetic_datanator.datanator.FormattedData`): list of reactions and their kinetic data
        output_filename (:obj:`str`): filename to store the list reaction kinetic data
    """

    # create a workbook
    wb = Workbook()

    # write kinetic data to a worksheet
    ws = wb.active
    ws.title = "Kinetics"

    headers = [
        'Name', 'Sabio Reaction IDs',
        'Vmax Median Value', 'Vmax Median Entry', 'Vmax Min Entry', 'Vmax Max Entry', 'Vmax Proximity', 'Vmax Lift Info' 'Closest Vmax', 'Sabio Entry IDs', 'Closest Vmax Values',
        'Km Median Value', 'Km Median Entry', 'Km Min Enry', 'Km Max Entry', 'Km Proximity', 'Km Lift Info', 'Closest Km Sabio Entry IDs', 'Closest Km Values'
    ]
    for i_col, value in enumerate(headers):
        ws.cell(row=1, column=i_col + 1).value = value

    for i_row, rxn in enumerate(reactions):
        ws.cell(row=i_row + 2, column=1).value = rxn.id
        ws.cell(row=i_row + 2, column=2).value = _format_list(rxn.reaction_ids)

        # vmax information
        ws.cell(row=i_row + 2, column=3).value = _format_object(rxn.vmax_data.median_entry, "vmax")
        ws.cell(row=i_row + 2, column=4).value = _format_object(rxn.vmax_data.median_entry)
        ws.cell(row=i_row + 2, column=5).value = _format_object(rxn.vmax_data.min_entry)
        ws.cell(row=i_row + 2, column=6).value = _format_object(rxn.vmax_data.max_entry)
        ws.cell(row=i_row + 2, column=7).value = _format_proximity(rxn.vmax_data.closest_entries)
        ws.cell(row=i_row + 2, column=8).value = _format_object(rxn.vmax_data, "lift_info")
        ws.cell(row=i_row + 2, column=9).value = _format_list(rxn.vmax_data.closest_entry_ids)
        ws.cell(row=i_row + 2, column=10).value = _format_list(rxn.vmax_data.closest_values)

        # km information
        ws.cell(row=i_row + 2, column=11).value = _format_object(rxn.km_data.median_entry, "km")
        ws.cell(row=i_row + 2, column=12).value = _format_object(rxn.km_data.median_entry)
        ws.cell(row=i_row + 2, column=13).value = _format_object(rxn.km_data.min_entry)
        ws.cell(row=i_row + 2, column=14).value = _format_object(rxn.km_data.max_entry)
        ws.cell(row=i_row + 2, column=15).value = _format_proximity(rxn.km_data.closest_entries)
        ws.cell(row=i_row + 2, column=16).value = _format_object(rxn.km_data, "lift_info")
        ws.cell(row=i_row + 2, column=17).value = _format_list(rxn.km_data.closest_entry_ids)
        ws.cell(row=i_row + 2, column=19).value = _format_list(rxn.km_data.closest_values)

    # write taxonomic information to a sheet
    ws = wb.create_sheet(title="Taxonomic information")
    ws.cell(row=1, column=1, value='Taxonomy')
    ws.cell(row=1, column=2, value='Proximity')
    for i_row, value in enumerate(taxon_finder.get_taxonomic_lineage(species)):
        ws.cell(row=i_row+2, column=1, value=value)
        ws.cell(row=i_row+2, column=2, value=i_row + 1)

    # save workbook to a file
    wb.save(output_filename)


def _format_object(obj, attr=None):
    """ Generate a string representation of an object or of an attribute of an object

    Args:
        obj (:obj:`object`): object
        attr (:obj:`str`, optional): attribute to format

    Returns:
        :obj:`str`: string representation of the object or of the attribute of the object
    """
    if obj is not None:
        if attr:
            return str(getattr(obj, attr))
        else:
            return str(obj.__dict__)

    return ""


def _format_proximity(entries):
    """ Generate a string representation of the proximity of the first of a list of entry

    Args:
        entries (:obj:`list of :obj:`kinetic_datanator.sabio_interface.Entry`): list of entries

    Returns:
        :obj:`str`: string representation of the proximity of the first of a list of entry
    """
    if entries:
        return str(entries[0].proximity)
    else:
        return ""


def _format_list(lst):
    """ Generate a comma-separated string representation of a list of values

    Args:
        lst (:obj:`list`): list of values

    Returns:
        :obj:`str`: comma-separated string representation of a list of values
    """
    return ', '.join(str(el) for el in lst)
