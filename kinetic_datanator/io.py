""" Utilities for reading and writing data to/from files

:Author: Yosef Roth <yosefdroth@gmail.com>
:Author: Jonathan Karr <jonrkarr@gmail.com>
:Date: 2017-04-14
:Copyright: 2017, Karr Lab
:License: MIT
"""

from kinetic_datanator.core import observation
from wc_utils.workbook.core import Row, Workbook, Worksheet
from wc_utils.workbook.io import WorkbookStyle, WorksheetStyle, write
import re
import openpyxl


class InputReader(object):

    def __init__(self):
        pass

    def run(self, filename):
        """ Read input data from an Excel workbook

        Args:
            filename (:obj:`str`): filename of Excel workbook

        Returns:
            :obj:`tuple`: 

                * :obj:`observation.Genetics`: genetics
                * :obj:`list` of :obj:`observation.Compartment`: list of compartments
                * :obj:`list` of :obj:`observation.Specie`: list of species
                * :obj:`list` of :obj:`observation.Reaction`: list of reactions
        """
        wb = openpyxl.load_workbook(filename=filename)
        genetics = self.read_genetics(wb.get_sheet_by_name('Genetics'))
        compartments = self.read_compartments(wb.get_sheet_by_name('Compartments'))
        species = self.read_species(wb.get_sheet_by_name('Species'))
        reactions = self.read_reactions(wb.get_sheet_by_name('Reactions'), compartments, species)

        return (genetics, compartments, species, reactions)

    def read_genetics(self, ws):
        """ Read taxon from an Excel worksheet

        Args:
            ws (:obj:`openpyxl.Worksheet`): worksheet

        Returns:
            :obj:`observation.Genetics`: taxon
        """
        return observation.Genetics(
                taxon=ws.cell(row=2, column=1).value,
                variation=ws.cell(row=2, column=2).value,
                )

    def read_compartments(self, ws):
        """ Read compartments from an Excel worksheet

        Args:
            ws (:obj:`openpyxl.Worksheet`): worksheet

        Returns:
            :obj:`list` of `observation.Compartment`: list of compartments
        """
        compartments = []
        for i in range(2, ws.max_row + 1):
            compartments.append(observation.Compartment(
                id=ws.cell(row=i, column=1).value,
                name=ws.cell(row=i, column=2).value,
            ))
        return compartments

    def read_species(self, ws):
        """ Read species from an Excel worksheet

        Args:
            ws (:obj:`openpyxl.Worksheet`): worksheet

        Returns:
            :obj:`list` of `observation.Specie`: list of species
        """
        species = []
        for i in range(2, ws.max_row + 1):
            species.append(observation.Specie(
                id=ws.cell(row=i, column=1).value,
                structure=ws.cell(row=i, column=2).value,
            ))

        return species

    def read_reactions(self, ws, compartments, species):
        """ Read reactions from an Excel worksheet

        Args:
            ws (:obj:`openpyxl.Worksheet`): worksheet
            compartments (:obj:`list` of :obj:`observation.Compartment`): list of compartments
            species (:obj:`list` of :obj:`observation.Specie`): list of species

        Returns:
            :obj:`list` of `observation.Reaction`: list of reactions
        """
        reactions = []
        for i in range(2, ws.max_row + 1):
            rxn = self.parse_reaction_equation(ws.cell(row=i, column=2).value, compartments, species)
            rxn.id = ws.cell(row=i, column=1).value
            reactions.append(rxn)
        return reactions

    def parse_reaction_equation(self, equation, compartments, species):
        """ Parse a reaction equation, e.g.

        * [c]: ATP + H2O ==> ADP + PI + H
        * GLC[e] + ATP[c] + H2O[c] ==> GLC[c] + ADP[c] + PI[c] + H[c]

        Args:
            equation (:obj:`str`): reaction equation
            compartments (:obj:`list` of :obj:`observation.Compartment`): list of compartments
            species (:obj:`list` of :obj:`observation.Specie`): list of species

        Returns:
            :obj:`observation.Reaction': reaction
        """
        compartments_dict = {c.id: c for c in compartments}
        species_dict = {s.id: s for s in species}

        global_comp = '\[(?P<comp>[a-z0-9_]+)\]: *'
        global_part = ' *(([0-9\.]+) )?([a-z0-9_]+)'
        global_lhs = '(?P<lhs>{}( *\+ *{})*)'.format(global_part, global_part)
        global_rhs = '(?P<rhs>{}( *\+ *{})*)'.format(global_part, global_part)
        local_part = ' *(([0-9\.]+) )?([a-z0-9_]+)\[([a-z0-9_]+)\]'
        local_lhs = '(?P<lhs>{}( *\+ *{})*)'.format(local_part, local_part)
        local_rhs = '(?P<rhs>{}( *\+ *{})*)'.format(local_part, local_part)
        sep = ' *(?P<sep><{0,1})[-=]{1,2}> *'

        global_pattern = global_comp + global_lhs + sep + global_rhs
        local_pattern = local_lhs + sep + local_rhs

        global_match = re.match(global_pattern, equation.rstrip(), re.IGNORECASE)
        local_match = re.match(local_pattern, equation.rstrip(), re.IGNORECASE)

        if global_match:
            global_comp = global_match.groupdict()['comp']
            sep = global_match.groupdict()['sep']
            lhs = global_match.groupdict()['lhs']
            rhs = global_match.groupdict()['rhs']

            participants = []
            for part in re.findall(global_part, lhs, re.IGNORECASE):
                participants.append(observation.ReactionParticipant(
                    specie=species_dict[part[2]],
                    compartment=compartments_dict[global_comp],
                    coefficient=-float(part[0][0:-1] or 1.),
                ))
            for part in re.findall(global_part, rhs, re.IGNORECASE):
                participants.append(observation.ReactionParticipant(
                    specie=species_dict[part[2]],
                    compartment=compartments_dict[global_comp],
                    coefficient=float(part[0][0:-1] or 1.),
                ))

        elif local_match:
            sep = local_match.groupdict()['sep']
            lhs = local_match.groupdict()['lhs']
            rhs = local_match.groupdict()['rhs']

            participants = []
            for part in re.findall(local_part, lhs, re.IGNORECASE):
                participants.append(observation.ReactionParticipant(
                    specie=species_dict[part[2]],
                    compartment=compartments_dict[part[3]],
                    coefficient=-float(part[0][0:-1] or 1.),
                ))
            for part in re.findall(local_part, rhs, re.IGNORECASE):
                participants.append(observation.ReactionParticipant(
                    specie=species_dict[part[2]],
                    compartment=compartments_dict[part[3]],
                    coefficient=float(part[0][0:-1] or 1.),
                ))
        else:
            raise ValueError('Reaction is not parseable: {}'.format(equation))

        # save specified participant order
        for i_part, part in enumerate(participants):
            part.order = i_part

        return observation.Reaction(participants=participants, reversible=sep == '<')


class ResultsWriter(object):
    """ Save reaction kinetic data to an Excel workbook """

    @classmethod
    def run(cls, taxon, reactions, filename):
        """ Save a list of reaction kinetic data to an Excel workbook or set of csv/tsv files

        Args:
            taxon (:obj:`str`): taxon
            reactions (:obj:`list` of :obj:`kinetic_datanator.datanator.SabioResult`): list of reactions and their kinetic data
            filename (:obj:`str`): filename to store the list reaction kinetic data
        """
        wb = Workbook()
        style = WorkbookStyle()

        # taxon
        style['Genetics'] = WorksheetStyle(head_rows=1, head_columns=0,
                                        head_row_font_bold=True, head_row_fill_fgcolor='CCCCCC', row_height=15)

        ws = wb['Genetics'] = Worksheet()
        ws.append(Row(['Name', 'NCBI ID']))
        ws.append(Row([taxon, taxonomy_util.Taxon(taxon).get_ncbi_id()]))

        # kinetics
        style['Kinetics'] = WorksheetStyle(head_rows=1, head_columns=0,
                                           head_row_font_bold=True, head_row_fill_fgcolor='CCCCCC', row_height=15)

        ws = wb['Kinetics'] = Worksheet()
        ws.append(Row([
            'Name', 'Sabio Reaction IDs',
            'Vmax Median Value', 'Vmax Median Entry', 'Vmax Min Entry', 'Vmax Max Entry', 'Vmax Proximity', 'Vmax Lift Info' 'Closest Vmax', 'Sabio Entry IDs', 'Closest Vmax Values',
            'Km Median Value', 'Km Median Entry', 'Km Min Enry', 'Km Max Entry', 'Km Proximity', 'Km Lift Info', 'Closest Km Sabio Entry IDs', 'Closest Km Values'
        ]))
        for i_row, rxn in enumerate(reactions):
            ws.append(Row([
                rxn.id,
                cls._format_list(rxn.reaction_ids),

                # vmax information
                cls._format_object(rxn.vmax_data.median_entry, "vmax"),
                cls._format_object(rxn.vmax_data.median_entry),
                cls._format_object(rxn.vmax_data.min_entry),
                cls._format_object(rxn.vmax_data.max_entry),
                cls._format_proximity(rxn.vmax_data.closest_entries),
                cls._format_object(rxn.vmax_data, "lift_info"),
                cls._format_list(rxn.vmax_data.closest_entry_ids),
                cls._format_list(rxn.vmax_data.closest_values),

                # km information
                cls._format_object(rxn.km_data.median_entry, "km"),
                cls._format_object(rxn.km_data.median_entry),
                cls._format_object(rxn.km_data.min_entry),
                cls._format_object(rxn.km_data.max_entry),
                cls._format_proximity(rxn.km_data.closest_entries),
                cls._format_object(rxn.km_data, "lift_info"),
                cls._format_list(rxn.km_data.closest_entry_ids),
                cls._format_list(rxn.km_data.closest_values),
            ]))

        # save to file
        write(filename, wb, style=style)

    @classmethod
    def run2(cls, taxon, reactions, filename):
        """ Save a list of reaction kinetic data to an Excel workbook or set of csv/tsv files

        Args:
            taxon (:obj:`str`): taxon name
            reactions (:obj:`list` of :obj:`kinetic_datanator.datanator.SabioResult`): list of reactions and their kinetic data
            filename (:obj:`str`): filename to store the list reaction kinetic data
        """
        # kinetics
        style['Kinetics'] = WorksheetStyle(head_rows=2, head_columns=0,
                                           head_row_font_bold=True, head_row_fill_fgcolor='CCCCCC', row_height=15)

        ws = wb['Kinetics'] = Worksheet()
        ws.append(Row([
            '',
            'Vmax', '', '', '', '', '' '', '', '',
            'Km', '', '', '', '', '' '', '', '',
            'Cross references', '', '',
        ]))
        ws.append(Row([
            'ID',
            'Median value', 'Median entry', 'Proximity', 'Lift Info' 'Closest Vmax', 'Min', 'Max', 'Sabio Entry IDs', 'Closest Values',
            'Median', 'Median Entry', 'Proximity', 'Lift Info' 'Closest Vmax', 'Min', 'Max', 'Sabio Entry IDs', 'Closest Values',
            'EC number', 'Predicted EC number', 'SABIO IDs',
        ]))

        for i_row, rxn in enumerate(reactions):
            row = [
                rxn.id,

                # vmax
                cls._format_object(rxn.vmax_data.median_entry, "vmax"),
                cls._format_object(rxn.vmax_data.median_entry),
                cls._format_object(rxn.vmax_data.min_entry),
                cls._format_object(rxn.vmax_data.max_entry),
                cls._format_proximity(rxn.vmax_data.closest_entries),
                cls._format_object(rxn.vmax_data, "lift_info"),
                cls._format_list(rxn.vmax_data.closest_entry_ids),
                cls._format_list(rxn.vmax_data.closest_values),

                # km
                cls._format_object(rxn.km_data.median_entry, "km"),
                cls._format_object(rxn.km_data.median_entry),
                cls._format_object(rxn.km_data.min_entry),
                cls._format_object(rxn.km_data.max_entry),
                cls._format_proximity(rxn.km_data.closest_entries),
                cls._format_object(rxn.km_data, "lift_info"),
                cls._format_list(rxn.km_data.closest_entry_ids),
                cls._format_list(rxn.km_data.closest_values),

                # cross references
                rxn.ec_number,
                cls._format_list(rxn.predicted_ec_numbers),
                cls._format_list(rxn.reaction_ids),
            ]

            ws.append(Row(row))

    @classmethod
    def _format_object(cls, obj, attr=None):
        """ Generate a string representation of an object or of an attribute of an object

        Args:
            obj (:obj:`object`): object
            attr (:obj:`str`, optional): attribute to format

        Returns:
            :obj:`str`: string representation of the object or of the attribute of the object
        """
        if obj is not None:
            if attr:
                return str(getattr(obj, attr))
            else:
                return str(obj.__dict__)

        return ""

    @classmethod
    def _format_proximity(cls, entries):
        """ Generate a string representation of the proximity of the first of a list of entry

        Args:
            entries (:obj:`list of :obj:`kinetic_datanator.sabio_rk.Entry`): list of entries

        Returns:
            :obj:`str`: string representation of the proximity of the first of a list of entry
        """
        if entries:
            return str(entries[0].proximity)
        else:
            return ""

    @classmethod
    def _format_list(cls, values):
        """ Generate a comma-separated string representation of a list of values

        Args:
            values (:obj:`list`): list of values

        Returns:
            :obj:`str`: comma-separated string representation of a list of values
        """
        return ', '.join(str(val) for val in values)
