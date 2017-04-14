""" 
:Author: Yosef Roth <yosefdroth@gmail.com>
:Date: 2017-04-06
:Copyright: 2017, Karr Lab
:License: MIT
"""

from . import data_structs
from . import inchi_generator
from . import io
from . import query_string_manipulator
import openpyxl




class ReactionQuery:
    def __init__(self, id):
        self.id = id
        self.substrates = [] #this is a list of compound objects
        self.products = [] #this is a list of compound objects
        self.num_participants = [] #this is an array of two numbers, first number for substrates, second for products
        self.keggID = ""
        self.ec_number = ""
        self.predicted_ec_number = ""

    def set_predicted_ec_number_from_ezyme_algorithm(self):        
        """ Predict EC number and store the value

        Returns:
            :obj:`str`: top-predicted EC number
        """

        from kinetic_datanator.util import molecule_util
        from kinetic_datanator.util import reaction_util

        parts = []

        for r in self.substrates:
            parts.append(reaction_util.ReactionParticipant(
                molecule=molecule_util.Molecule(r.inchi_smiles, name=r.id),
                coefficient=-1,
            ))

        for p in self.products:
            parts.append(reaction_util.ReactionParticipant(
                molecule=molecule_util.Molecule(p.inchi_smiles, name=p.id),
                coefficient=1,
            ))

        rxn = reaction_util.Reaction(parts)
        ec_numbers = reaction_util.Ezyme.run(rxn)
        if ec_numbers:
            ec_number = ec_numbers[0].ec_number
        else:
            ec_number = ''


        self.predicted_ec_number = ec_number




def generate_reaction_queries(filename):
    #Input an excel sheet data object (from openpyxl)
    #It outputs a list of reaction_query Objects

    wb = openpyxl.load_workbook(filename=filename)
    

    #this gets a dict that is used to correlate inchiString with Sabio's name for that compount
    sabioNameToInchiDict = inchi_generator.getSabioNameToInchiDict()

    #this instantiates a compound object for each metabolite in the excel sheet
    compoundList = io.InputReader.read_compounds(wb)

    #this creates a dictionary of ID to compound to make it quick to find the compound object 
    #with the compound ID. This works because each compound ID must be unique
    idToCompound = {} #this will be a dict used to find compound information about each metab id
    for comp in compoundList:
        idToCompound[comp.id] = comp


    #this creates a dictionary of reaction ID to reaction string
    #to work on: making it possible to only enter one row
    #to work on: make it possible not to have any reaction string data to begin with
    ws = wb.get_sheet_by_name('Reactions')
    reactions = []
    for i in range(2, ws.max_row + 1):
        reactions.append({
        	'id': ws.cell(row=i, column=1).value, 
        	'stoichiometry': io.InputReader.parse_reaction_stoichiometry(ws.cell(row=i, column=2).value),
        	})


    #this is is where the ReactionQuery objects are instantiated
    #each reaction is parsed, each metabolite ID is converted to a Compound object
    #then the fields for each ReactionQuery object are set
    #finally the ReactionQuery object is added to the reaction_queries list
    reaction_queries = [] #this is the only output of this method. Its a list of reaction_queries
    for rxn in reactions:
        query = ReactionQuery(rxn['id'])
        query.num_participants = [len(rxn['stoichiometry'][0]), len(rxn['stoichiometry'][1])]

        compoundBalancedMetab = []
        for side in rxn['stoichiometry']:
            compoundSide = []
            for metab in side:
                try:
                    compoundSide.append(idToCompound[metab])
                except:
                    newUnrecognizedCompound = data_structs.Compound(metab)
                    compoundSide.append(newUnrecognizedCompound)
            compoundBalancedMetab.append(compoundSide)
        query.substrates = compoundBalancedMetab[0]
        query.products = compoundBalancedMetab[1]
        reaction_queries.append(query)

    return reaction_queries

