#import openpyxl
from openpyxl import Workbook
import os
import TaxonFinder


#takes in FormattedData as an input. Saves a file 

def createExcelSheet(outputFilename, FormattedDataList, species):
	topRow = [['Name', 'Sabio Reaction IDs', 
	'Vmax Median Value', 'Vmax Median Entry', 'Vmax Min Entry', 'Vmax Max Entry', 'Vmax Proximity', 'Vmax Lift Info' 'Closest Vmax', 'Sabio Entry IDs', 'Closest Vmax Values',
	'Km Median Value', 'Km Median Entry', 'Km Min Enry', 'Km Max Entry', 'Km Proximity', 'Km Lift Info', 'Closest Km Sabio Entry IDs', 'Closest Km Values']]

	rows = []
	for formattedData in FormattedDataList:
		row = []
		#first generation reaction information
		row = [formattedData.id, toCSV(formattedData.reactionIDs),

		#then vmax information
		formatValue(formattedData.VmaxData.medianEntry, "vmax"), formatValue(formattedData.VmaxData.medianEntry), formatValue(formattedData.VmaxData.minEntry), formatValue(formattedData.VmaxData.maxEntry),
		formatProxim(formattedData.VmaxData.closestEntries), formatValue(formattedData.VmaxData, "liftInfo"), toCSV(formattedData.VmaxData.closestEntryIDs), toCSV(formattedData.VmaxData.closestValues),

		#finally, the km information
		formatValue(formattedData.KmData.medianEntry, "km"), formatValue(formattedData.KmData.medianEntry), formatValue(formattedData.KmData.minEntry), formatValue(formattedData.KmData.maxEntry),
		formatProxim(formattedData.KmData.closestEntries), formatValue(formattedData.KmData, "liftInfo"), toCSV(formattedData.KmData.closestEntryIDs), toCSV(formattedData.KmData.closestValues)]
		rows.append(row)


	rows = topRow + rows

	wb = Workbook()
	dest_filename = outputFilename
	ws1 = wb.active
	ws1.title = "Kinetics"


	row = 1
	while row < len(rows)+1:
		col = 1
		while col < len(rows[0])+1:
			_ = ws1.cell(column=col, row=row, value=rows[row-1][col-1])
			col = col+1
		row = row+1

	ws2 = wb.create_sheet(title="Taxnomic Information")
	#ws2.title = "Taxnomic Information"

	columns = []
	leftColumn = []
	leftColumn.append("Taxonomy")
	leftColumn = leftColumn+TaxonFinder.getTaxonomicLineage(species)
	rightColumn = []
	rightColumn.append("Proximity")
	rightColumn = rightColumn + range(1, len(leftColumn))
	columns.append(rightColumn)
	columns.append(leftColumn)
	
	row = 1
	while row < len(columns[0])+1:
		col = 1
		while col < len(columns)+1:
			_ = ws2.cell(column=col, row=row, value=columns[col-1][row-1])
			col = col+1
		row = row+1


	wb.save(os.path.join('.', dest_filename))

def formatValue(outerField, innerField = ""):
	value = ""
	if outerField != None:
		if innerField:
			value =  "{}".format(outerField.__dict__[innerField])
		else:
			value = "{}".format(outerField.__dict__)
	if outerField == None:
		value =  "No Data"
	return value

def formatProxim(closestEntries):
	if len(closestEntries)>0:
		return "{}".format(closestEntries[0].proximity)
	else:
		return "No Data"

def toCSV(array):

	csv = ""
	for item in array:
		csv = csv + "{}".format(item) + ", "

	if len(csv)>0:
		csv = csv[:-2]
	return csv


if __name__ == '__main__':
	blue = None
	print(blue)